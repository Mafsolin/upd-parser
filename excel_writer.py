"""
excel_writer.py — запись извлечённых данных из УПД в Excel-файл.
"""

import logging
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE
from data_normalizer import excel_numeric_value, normalize_document

logger = logging.getLogger(__name__)


def _safe_excel_text(value: str) -> str:
    """Force untrusted OCR text to remain a string in spreadsheet programs."""
    return "'" + value if value.startswith(("=", "+", "-", "@")) else value

# ────────────────────────────────────────────────────────────────────────────
# Константы
# ────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "№",
    "Дата",
    "Наименование",
    "Ед.изм",
    "Количество",
    "Цена",
    "Сумма с НДС",
    "НДС",
    "№ счета фактуры",
    "Поставщик",
]

# Ширина колонок (в символах)
COLUMN_WIDTHS = [6, 14, 50, 10, 12, 14, 14, 14, 22, 40]

# Цвет шапки
HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
DATA_FONT    = Font(color="FF000000", size=10)


# ────────────────────────────────────────────────────────────────────────────
# Вспомогательные функции
# ────────────────────────────────────────────────────────────────────────────

def _create_workbook(path: Path) -> Workbook:
    """Создаёт новый Excel-файл с заголовками."""
    wb = Workbook()
    ws = wb.active
    ws.title = "УПД"

    # Заголовки
    ws.append(HEADERS)

    # Стили шапки
    for col_idx, (cell, width) in enumerate(
        zip(ws[1], COLUMN_WIDTHS), start=1
    ):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"       # шапка всегда видна при прокрутке

    wb.save(path)
    logger.info("Создан новый Excel-файл: %s", path)
    return wb


def _load_or_create(path: Path) -> tuple[Workbook, object]:
    """Открывает существующий файл или создаёт новый. Возвращает (wb, ws)."""
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        logger.debug("Открыт существующий файл: %s (%d строк данных)", path, ws.max_row - 1)
    else:
        wb = _create_workbook(path)
        ws = wb.active
    return wb, ws


def _last_row_number(ws) -> int:
    """Возвращает максимальный порядковый № в первой колонке (для авто-нумерации)."""
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, int) and val > max_num:
            max_num = val
    return max_num


# ────────────────────────────────────────────────────────────────────────────
# Основной класс
# ────────────────────────────────────────────────────────────────────────────

class ExcelWriter:
    """Записывает данные одного УПД (несколько товарных строк) в Excel."""

    def __init__(self, file_path: Path | str | None = None):
        """
        Args:
            file_path: путь к Excel-файлу. Если None — использует EXCEL_FILE из config.
        """
        self.path = Path(file_path) if file_path else Path(EXCEL_FILE)

    def write(self, data: dict, doc_name: str) -> int:
        """
        Записывает товары из `data` в Excel.

        Args:
            data:     словарь с полями date, seller, invoice_number, items[].
            doc_name: имя папки документа (для логов).

        Returns:
            Количество добавленных строк.
        """
        data = normalize_document(data)

        wb, ws = _load_or_create(self.path)
        counter     = _last_row_number(ws)
        rows_added  = 0

        date           = _safe_excel_text(data.get("date", ""))
        seller         = _safe_excel_text(data.get("seller", ""))
        invoice_number = _safe_excel_text(data.get("invoice_number", ""))
        items          = data.get("items", [])

        if not items:
            logger.warning("[%s] Список товаров пуст — нечего записывать.", doc_name)
            wb.close()
            return 0

        for item in items:
            counter    += 1
            rows_added += 1
            row = [
                counter,
                date,
                _safe_excel_text(item.get("name",  "")),
                _safe_excel_text(item.get("unit",  "")),
                excel_numeric_value(item.get("qty", "")),
                excel_numeric_value(item.get("price", "")),
                excel_numeric_value(item.get("total_with_vat", item.get("cost", ""))),
                excel_numeric_value(item.get("tax", ""), allow_without_vat=True),
                invoice_number,
                seller,
            ]
            ws.append(row)

            # Стили строки данных
            last_row = ws.max_row
            for col_idx, cell in enumerate(ws[last_row], start=1):
                cell.font      = DATA_FONT
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 3 else "center",
                    vertical="center",
                    wrap_text=True,
                )
                if col_idx in {5, 6, 7, 8} and isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = "0.##############"
            ws.row_dimensions[last_row].height = 18

        wb.save(self.path)
        wb.close()

        logger.info(
            "[%s] Записано строк в Excel: %d (файл: %s)",
            doc_name, rows_added, self.path,
        )
        return rows_added

    # ── приватные ────────────────────────────────────────────────────────────

    @staticmethod
    def _validate(data: dict) -> None:
        """Базовая проверка структуры данных от LLM."""
        if not isinstance(data, dict):
            raise ValueError(f"Ожидался dict, получен {type(data).__name__}.")
        if "items" not in data:
            raise ValueError("Поле 'items' отсутствует в ответе LLM.")
        if not isinstance(data["items"], list):
            raise ValueError("Поле 'items' должно быть списком.")
