import tempfile
import unittest
import os
from pathlib import Path
from unittest import mock

from excel_writer import ExcelWriter, HEADERS


class GuiSequentialProcessingTests(unittest.TestCase):
    def test_excel_headers_include_vat_included_sum(self):
        self.assertIn("Сумма с НДС", HEADERS)
        self.assertIn("Количество", HEADERS)
        self.assertIn("Поставщик", HEADERS)

    def test_excel_writer_uses_total_with_vat_and_keeps_missing_supplier_empty(self):
        data = {
            "date": "",
            "seller": "",
            "invoice_number": "",
            "items": [
                {
                    "name": "Перчатки",
                    "qty": "10",
                    "total_with_vat": "1200.00",
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "result.xlsx"
            rows = ExcelWriter(file_path=path).write(data, "img001")

            from openpyxl import load_workbook

            wb = load_workbook(path)
            ws = wb.active
            values = [cell.value for cell in ws[2]]
            wb.close()

        self.assertEqual(rows, 1)
        self.assertEqual(values[2], "Перчатки")
        self.assertEqual(values[4], 10)
        self.assertEqual(values[6], 1200)
        self.assertIsInstance(values[6], (int, float))
        self.assertIn(values[8], ("", None))
        self.assertIn(values[9], ("", None))

    def test_sequence_processing_treats_each_image_as_separate_document(self):
        from standalone_upd import process_upd

        calls = []
        writes = []

        class FakeParser:
            def parse_document(self, folder):
                image_names = sorted(path.name for path in Path(folder).iterdir())
                calls.append(image_names)
                return {
                    "date": "",
                    "seller": "",
                    "invoice_number": "",
                    "items": [{"name": image_names[0], "qty": "1", "total_with_vat": "10"}],
                }

        class FakeWriter:
            def __init__(self, file_path):
                self.file_path = file_path

            def write(self, data, doc_name):
                writes.append((doc_name, data["items"][0]["name"]))
                return 1

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            images = [root / "a.jpg", root / "b.jpg"]
            for image in images:
                image.write_bytes(b"x")

            summaries, rows = process_upd.process_image_sequence(
                images,
                root / "out.xlsx",
                parser_factory=FakeParser,
                writer_factory=FakeWriter,
            )

        self.assertEqual(calls, [["page001.jpg"], ["page001.jpg"]])
        self.assertEqual([summary["file"] for summary in summaries], ["a.jpg", "b.jpg"])
        self.assertEqual(rows, 2)
        self.assertEqual(len(writes), 2)

    def test_gui_save_dialog_returns_selected_xlsx_path_without_creating_output_folder(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            selected = root / "custom_report"
            output_dir = root / "output"

            with mock.patch(
                "tkinter.filedialog.asksaveasfilename",
                return_value=str(selected),
            ) as ask_save:
                with mock.patch.object(process_upd, "APP_DIR", root):
                    path = process_upd.ask_excel_output_path(parent=None)

            self.assertFalse(output_dir.exists())

        self.assertEqual(path, selected.with_suffix(".xlsx"))
        kwargs = ask_save.call_args.kwargs
        self.assertEqual(kwargs["defaultextension"], ".xlsx")
        self.assertEqual(kwargs["initialdir"], str(root))
        self.assertTrue(kwargs["initialfile"].startswith("upd_result_"))

    def test_gui_save_dialog_cancel_returns_none(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch("tkinter.filedialog.asksaveasfilename", return_value=""):
                path = process_upd.ask_excel_output_path(
                    parent=None,
                    initial_dir=Path(tmp),
                )

        self.assertIsNone(path)

    def test_existing_selected_excel_file_is_preserved_until_atomic_commit(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            path.write_text("old", encoding="utf-8")

            prepared = process_upd.prepare_selected_excel_path(path)
            staging = process_upd.build_staging_excel_path(prepared)
            staging.write_text("new", encoding="utf-8")

            self.assertEqual(prepared, path)
            self.assertEqual(path.read_text(encoding="utf-8"), "old")
            process_upd.commit_staged_excel(staging, prepared)
            self.assertEqual(path.read_text(encoding="utf-8"), "new")
            self.assertFalse(staging.exists())

    def test_selected_output_is_always_xlsx(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            prepared = process_upd.prepare_selected_excel_path(Path(tmp) / "report.txt")
            self.assertEqual(prepared.suffix, ".xlsx")

    def test_failed_staging_does_not_touch_existing_excel(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            path.write_text("old", encoding="utf-8")
            staging = process_upd.build_staging_excel_path(path)
            process_upd.discard_staged_excel(staging)
            self.assertEqual(path.read_text(encoding="utf-8"), "old")

    def test_close_is_blocked_while_processing(self):
        from standalone_upd import process_upd

        app = process_upd.MinimalUpdApp.__new__(process_upd.MinimalUpdApp)
        app.processing = True
        app.language = "en"
        app.root = mock.Mock()
        with mock.patch("tkinter.messagebox.showwarning") as warning:
            app.on_close()
        app.root.destroy.assert_not_called()
        warning.assert_called_once()

        app.processing = False
        app.on_close()
        app.root.destroy.assert_called_once()

    def test_settings_without_provider_requires_configuration(self):
        from standalone_upd import process_upd

        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(RuntimeError, "Провайдер не настроен"):
                process_upd.ensure_api_key(Path(tmp), prompt_fn=None)

    def test_text_context_menu_installs_right_click_paste(self):
        from standalone_upd import process_upd

        class FakeMenu:
            def __init__(self):
                self.commands = []
                self.posted_at = None

            def add_command(self, label, command):
                self.commands.append((label, command))

            def tk_popup(self, x, y):
                self.posted_at = (x, y)

            def grab_release(self):
                pass

        class FakeWidget:
            def __init__(self):
                self.bindings = {}
                self.events = []
                self.inserted = []

            def bind(self, sequence, callback):
                self.bindings[sequence] = callback

            def event_generate(self, sequence):
                self.events.append(sequence)

            def clipboard_get(self):
                return "api-key-from-clipboard"

            def delete(self, _start, _end):
                pass

            def insert(self, _index, value):
                self.inserted.append(value)

        menu = FakeMenu()
        widget = FakeWidget()

        returned_menu = process_upd.install_text_context_menu(
            widget,
            menu_factory=lambda _widget: menu,
        )

        self.assertIs(returned_menu, menu)
        self.assertIn("<Button-3>", widget.bindings)
        labels = [label for label, _command in menu.commands]
        self.assertIn("Вставить", labels)
        paste_command = dict(menu.commands)["Вставить"]
        paste_command()
        self.assertEqual(widget.inserted, ["api-key-from-clipboard"])
        self.assertIn("<Control-KeyPress>", widget.bindings)
        event = type("Event", (), {"keycode": 86})()
        self.assertEqual(widget.bindings["<Control-KeyPress>"](event), "break")
        for keycode, sequence in ((65, "<<SelectAll>>"), (67, "<<Copy>>"), (88, "<<Cut>>")):
            with self.subTest(keycode=keycode):
                result = widget.bindings["<Control-KeyPress>"](type("Event", (), {"keycode": keycode})())
                self.assertEqual(result, "break")
                self.assertIn(sequence, widget.events)
        self.assertIn("<Shift-Insert>", widget.bindings)


if __name__ == "__main__":
    unittest.main()
