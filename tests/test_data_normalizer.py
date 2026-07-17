import tempfile
import unittest
from pathlib import Path

from data_normalizer import excel_numeric_value, normalize_document, normalize_numeric_value
from excel_writer import ExcelWriter


class DataNormalizerTests(unittest.TestCase):
    def test_numeric_values_use_comma_without_spaces_or_minus(self):
        self.assertEqual(normalize_numeric_value(" 1 234.50- "), "1234,50")
        self.assertEqual(normalize_numeric_value("−10.25"), "10,25")
        self.assertEqual(normalize_numeric_value("1\u00a0234.5"), "1234,5")

    def test_excel_numeric_value_is_a_number_not_text(self):
        self.assertEqual(excel_numeric_value("1 234,50-"), 1234.50)
        self.assertEqual(excel_numeric_value("без НДС", allow_without_vat=True), "без НДС")

    def test_normalizes_only_exported_numeric_fields(self):
        data = normalize_document({
            "date": "01.02.2026",
            "seller": "ООО Точка-Экспорт",
            "invoice_number": "СФ-1.2",
            "items": [{
                "name": "Товар-1.2",
                "unit": "шт.",
                "qty": "1.50-",
                "price": "2 000.75-",
                "total_with_vat": "3.25-",
                "tax": "0.50-",
            }],
        })
        self.assertEqual(data["seller"], "ООО Точка-Экспорт")
        self.assertEqual(data["invoice_number"], "СФ-1.2")
        self.assertEqual(data["items"][0]["name"], "Товар-1.2")
        self.assertEqual(
            [data["items"][0][field] for field in ("qty", "price", "total_with_vat", "tax")],
            ["1,50", "2000,75", "3,25", "0,50"],
        )

    def test_excel_writer_applies_normalization_to_all_numeric_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "result.xlsx"
            ExcelWriter(path).write({"items": [{
                "name": "Товар", "qty": "1.1-", "price": "2.2-",
                "total_with_vat": "3.3-", "tax": "4.4-",
            }]}, "doc")
            from openpyxl import load_workbook
            workbook = load_workbook(path)
            values = [workbook.active.cell(2, column).value for column in (5, 6, 7, 8)]
            workbook.close()
        self.assertEqual(values, [1.1, 2.2, 3.3, 4.4])

    def test_rejects_invalid_item_shape(self):
        with self.assertRaisesRegex(ValueError, "Товарная строка"):
            normalize_document({"items": ["not an object"]})


if __name__ == "__main__":
    unittest.main()
